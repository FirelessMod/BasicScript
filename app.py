import openpyxl as xl
file = input("What is the name of the file? \n")
wb = xl.load_workbook(f'{file}.xlsx')
sheet = wb['Sheet1']
correct_name = sheet.cell(1, 4)
correct_name.value = "updated price"

for row in range(2,    sheet.max_row + 1 ):
    cell = sheet.cell(row, 3)
    correct_price = cell.value * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_price

wb.save('transactions_corrected.xlsx')
